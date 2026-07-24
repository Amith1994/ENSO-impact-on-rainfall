import os
# Configure OpenBLAS and standard thread variables to 1 before imports to prevent memory allocation crashes in child processes
os.environ['OPENBLAS_NUM_THREADS'] = '1'
os.environ['MKL_NUM_THREADS'] = '1'
os.environ['OMP_NUM_THREADS'] = '1'
os.environ['NUMEXPR_NUM_THREADS'] = '1'

import matplotlib
matplotlib.use('Agg')  # Set non-interactive backend globally first

import sys
import logging
import argparse
import pandas as pd
import numpy as np
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# Ensure local modules can be found
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

# Configure Logging
log_file = os.path.join('Results', 'run_log.txt') if os.path.exists('Results') else 'run_log.txt'
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(log_file, mode='w', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

# Multiprocessing Worker Function
def process_district_worker(args):
    file_path, enso_map, output_dir = args
    
    try:
        from read_data import load_district_data
        from calculate_normals import compute_district_normals
        from departure import compute_district_departures
        from monthly import process_monthly_enso_summary
        from weekly import process_weekly_enso_summary
        from season import process_seasonal_enso_summary, process_annual_enso_summary
        from stats_analysis import run_group_tests, run_correlation_and_regression, run_mann_kendall_trend
        from graphs import (
            plot_monthly_bar_comparison, plot_monthly_anomaly_line,
            plot_monthly_boxplot_violin, plot_monthly_heatmap,
            plot_weekly_ribbon_anomaly, plot_seasonal_annual_boxplots
        )
        
        # 1. Load Data
        district_name, df = load_district_data(file_path)
        if df is None:
            return None
            
        # 2. Compute Normals
        normals = compute_district_normals(df)
        
        # 3. Compute Departures
        deps = compute_district_departures(df, normals, district_name)
        
        # Extracted tables
        monthly_dep = deps['monthly']
        weekly_dep = deps['weekly']
        seasonal_dep = deps['seasonal']
        annual_dep = deps['annual']
        
        # Add ENSO_Phase column to raw dataframes for plotting and statistical tests
        monthly_dep['ENSO_Phase'] = monthly_dep['Year'].map(enso_map)
        weekly_dep['ENSO_Phase'] = weekly_dep['Year'].map(enso_map)
        seasonal_dep['ENSO_Phase'] = seasonal_dep['Year'].map(enso_map)
        annual_dep['ENSO_Phase'] = annual_dep['Year'].map(enso_map)
        
        # 4. Generate Summaries by ENSO Phase
        m_summary = process_monthly_enso_summary(monthly_dep, enso_map)
        w_summary = process_weekly_enso_summary(weekly_dep, enso_map)
        s_summary = process_seasonal_enso_summary(seasonal_dep, enso_map)
        a_summary = process_annual_enso_summary(annual_dep, enso_map)
        
        # 5. Statistical Hypothesis & Trend Tests
        # Group differences (ANOVA / Kruskal-Wallis)
        anova_results = []
        for var_name, df_var in [('SW Monsoon', seasonal_dep[seasonal_dep['Season'] == 'SW Monsoon']),
                                 ('NE Monsoon', seasonal_dep[seasonal_dep['Season'] == 'NE Monsoon']),
                                 ('Annual', annual_dep)]:
            df_var_enso = df_var.copy()
            df_var_enso['ENSO_Phase'] = df_var_enso['Year'].map(enso_map)
            res_g = run_group_tests(df_var_enso, val_col='Actual', group_col='ENSO_Phase')
            anova_results.append({
                'District': district_name,
                'Variable': var_name,
                'ANOVA_F': res_g['anova_f'],
                'ANOVA_p': res_g['anova_p'],
                'KW_H': res_g['kw_h'],
                'KW_p': res_g['kw_p']
            })
            
        # Correlations and Regression
        corr_results = []
        for var_name, df_var in [('SW Monsoon', seasonal_dep[seasonal_dep['Season'] == 'SW Monsoon']),
                                 ('NE Monsoon', seasonal_dep[seasonal_dep['Season'] == 'NE Monsoon']),
                                 ('Annual', annual_dep)]:
            res_c = run_correlation_and_regression(df_var, enso_map, val_col='Actual')
            corr_results.append({
                'District': district_name,
                'Variable': var_name,
                **res_c
            })
            
        # Trend Analysis (Mann-Kendall & Sen's Slope)
        trend_results = []
        # Annual
        res_t = run_mann_kendall_trend(annual_dep['Year'].values, annual_dep['Actual'].values)
        trend_results.append({'District': district_name, 'Period': 'Annual', **res_t})
        # SW Monsoon
        df_sw = seasonal_dep[seasonal_dep['Season'] == 'SW Monsoon']
        res_t = run_mann_kendall_trend(df_sw['Year'].values, df_sw['Actual'].values)
        trend_results.append({'District': district_name, 'Period': 'SW Monsoon', **res_t})
        # NE Monsoon
        df_ne = seasonal_dep[seasonal_dep['Season'] == 'NE Monsoon']
        res_t = run_mann_kendall_trend(df_ne['Year'].values, df_ne['Actual'].values)
        trend_results.append({'District': district_name, 'Period': 'NE Monsoon', **res_t})
        # Monthly (1-12)
        for m in range(1, 13):
            df_m = monthly_dep[monthly_dep['Month'] == m]
            res_t = run_mann_kendall_trend(df_m['Year'].values, df_m['Actual'].values)
            trend_results.append({'District': district_name, 'Period': f"Month {m}", **res_t})
            
        # 6. Generate Figures
        fig_dir = os.path.join(output_dir, 'Figures')
        plot_monthly_bar_comparison(m_summary, district_name, fig_dir)
        plot_monthly_anomaly_line(monthly_dep, district_name, fig_dir)
        plot_monthly_boxplot_violin(monthly_dep, district_name, fig_dir)
        plot_monthly_heatmap(m_summary, district_name, fig_dir)
        plot_weekly_ribbon_anomaly(weekly_dep, district_name, fig_dir)
        plot_seasonal_annual_boxplots(seasonal_dep, annual_dep, district_name, fig_dir)
        
        return {
            'district_name': district_name,
            'monthly_dep': monthly_dep,
            'weekly_dep': weekly_dep,
            'seasonal_dep': seasonal_dep,
            'annual_dep': annual_dep,
            'm_summary': m_summary,
            'w_summary': w_summary,
            's_summary': s_summary,
            'a_summary': a_summary,
            'anova': pd.DataFrame(anova_results),
            'corr_reg': pd.DataFrame(corr_results),
            'trends': pd.DataFrame(trend_results)
        }
    except Exception as e:
        logger.error(f"Error processing file {file_path}: {e}", exc_info=True)
        return None

def write_styled_excel(writer, df, sheet_name):
    """
    Writes a DataFrame to an Excel worksheet with clean and professional styling.
    Uses bulk range operations instead of per-cell loops for speed.
    """
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    # Enable grid lines
    worksheet.views.sheetView[0].showGridLines = True

    # Styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2B547E', end_color='2B547E', fill_type='solid')
    cell_font = Font(name='Arial', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    right_align  = Alignment(horizontal='right',  vertical='center')
    left_align   = Alignment(horizontal='left',   vertical='center')

    thin_side = Side(style='thin', color='CCCCCC')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    n_cols = len(df.columns)
    n_rows = len(df)

    # --- Header row (row 1) ---
    for col_idx in range(1, n_cols + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # --- Detect numeric columns once ---
    float_cols = set()
    int_cols   = set()
    str_cols   = set()
    for col_idx, col in enumerate(df.columns, 1):
        series = df[col].dropna()
        if series.empty:
            continue
        if pd.api.types.is_float_dtype(series):
            float_cols.add(col_idx)
        elif pd.api.types.is_integer_dtype(series):
            int_cols.add(col_idx)
        else:
            str_cols.add(col_idx)

    # --- Data rows --- apply font + border in one pass; alignment/format per column type
    for col_idx in range(1, n_cols + 1):
        # Determine alignment & number format for this column
        if col_idx in float_cols:
            align  = right_align
            nfmt   = '0.00'
        elif col_idx in int_cols:
            align  = right_align
            nfmt   = None
        else:
            align  = left_align
            nfmt   = None

        for row_idx in range(2, n_rows + 2):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font      = cell_font
            cell.border    = thin_border
            cell.alignment = align
            if nfmt:
                cell.number_format = nfmt

    # --- Column widths: compute from dtype-aware max lengths ---
    for col_idx, col in enumerate(df.columns, 1):
        header_len = len(str(col))
        try:
            if col_idx in float_cols:
                max_val_len = df.iloc[:, col_idx - 1].map(lambda v: len(f"{v:.2f}") if pd.notna(v) else 0).max()
            else:
                max_val_len = df.iloc[:, col_idx - 1].astype(str).map(len).max()
        except Exception:
            max_val_len = 10
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = max(header_len, max_val_len, 10) + 3


def main():
    # 1. Initialize Directories
    output_dir = 'Results'
    sub_dirs = ['Monthly', 'Weekly', 'Seasonal', 'Annual', 'Statistics', 'Figures', 'Tables', 'Maps', 'Excel', 'Reports']
    for sub in sub_dirs:
        os.makedirs(os.path.join(output_dir, sub), exist_ok=True)
        
    # Relocate log file now that folders are created
    global log_file
    log_file_new = os.path.join(output_dir, 'Reports', 'run_log.txt')
    logging.getLogger().handlers[1].close()
    logging.getLogger().handlers[1] = logging.FileHandler(log_file_new, mode='w', encoding='utf-8')
    logger.info("Output directory structure initialized successfully.")
    
    # 2. Read ENSO Map
    try:
        from read_data import load_enso_classification, get_all_district_files
        enso_map = load_enso_classification('ENSO.txt')
    except Exception as e:
        logger.error(f"Failed to load ENSO classification: {e}")
        sys.exit(1)
        
    # 3. Get District Files
    district_files = get_all_district_files('Karnataka')
    if not district_files:
        logger.error("No district data files found in 'Karnataka/' directory.")
        sys.exit(1)
        
    logger.info(f"Found {len(district_files)} district files for analysis.")

    # 4. Process Districts in Parallel
    # Parse --full-quality flag: set env var before any graph worker can read it
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument('--full-quality', action='store_true', help='Save all 6 figure formats (PDF, SVG, TIFF, PNG) instead of fast 300dpi-PNG-only mode.')
    args, _ = parser.parse_known_args()
    if args.full_quality:
        os.environ['ENSO_SAVE_QUALITY'] = 'full'
        logger.info("Full-quality figure export enabled (PDF + SVG + PNG + TIFF at 300 & 600 dpi).")
    else:
        os.environ.setdefault('ENSO_SAVE_QUALITY', 'fast')
        logger.info("Fast figure export mode: saving 300 dpi PNG only. Run with --full-quality for publication output.")

    compiled_results = []
    worker_args = [(f, enso_map, output_dir) for f in district_files]

    logger.info("Spawning parallel district processing workers...")
    # Use all available CPU cores (one per district file, capped at cpu_count)
    n_workers = min(os.cpu_count() or 4, len(district_files))
    logger.info(f"Using {n_workers} parallel workers (cpu_count={os.cpu_count()}, districts={len(district_files)}).")
    with ProcessPoolExecutor(max_workers=n_workers) as executor:
        futures = {executor.submit(process_district_worker, args): args[0] for args in worker_args}
        for fut in as_completed(futures):
            f_path = futures[fut]
            res = fut.result()
            if res is not None:
                compiled_results.append(res)
                logger.info(f"Finished processing district: {res['district_name']}")
            else:
                logger.warning(f"District process returned None for file: {f_path}")
                
    if not compiled_results:
        logger.error("All district processing runs failed.")
        sys.exit(1)
        
    logger.info("Successfully compiled parallel results. Running aggregations...")
    
    # 5. Compile and Aggregate Combined Tables
    all_monthly_dep = pd.concat([r['monthly_dep'] for r in compiled_results], ignore_index=True)
    all_weekly_dep = pd.concat([r['weekly_dep'] for r in compiled_results], ignore_index=True)
    all_seasonal_dep = pd.concat([r['seasonal_dep'] for r in compiled_results], ignore_index=True)
    all_annual_dep = pd.concat([r['annual_dep'] for r in compiled_results], ignore_index=True)
    
    all_m_summary = pd.concat([r['m_summary'] for r in compiled_results], ignore_index=True)
    all_w_summary = pd.concat([r['w_summary'] for r in compiled_results], ignore_index=True)
    all_s_summary = pd.concat([r['s_summary'] for r in compiled_results], ignore_index=True)
    all_a_summary = pd.concat([r['a_summary'] for r in compiled_results], ignore_index=True)
    
    all_anova = pd.concat([r['anova'] for r in compiled_results], ignore_index=True)
    all_corr_reg = pd.concat([r['corr_reg'] for r in compiled_results], ignore_index=True)
    all_trends = pd.concat([r['trends'] for r in compiled_results], ignore_index=True)
    
    # Generate ENSO Summary Table combined
    # Create unified summary format: District, Season/Monthly, ENSO_Phase, Mean, SD, CV, Departure
    m_summary_formatted = all_m_summary.copy().rename(columns={'Month': 'Period'})
    m_summary_formatted['Period'] = m_summary_formatted['Period'].apply(lambda m: f"Month {m}")
    s_summary_formatted = all_s_summary.copy().rename(columns={'Season': 'Period'})
    a_summary_formatted = all_a_summary.copy()
    a_summary_formatted['Period'] = 'Annual'
    
    combined_enso_summary = pd.concat([
        s_summary_formatted,
        a_summary_formatted,
        m_summary_formatted
    ], ignore_index=True)
    # Order columns
    combined_enso_summary = combined_enso_summary[['District', 'Period', 'ENSO_Phase', 'Mean', 'Median', 'SD', 'CV', 'Max', 'Min', 'Departure']]
    
    # Save CSV Tables
    logger.info("Writing tabular outputs to CSV...")
    all_monthly_dep.to_csv(os.path.join(output_dir, 'Tables', 'district_monthly_departures.csv'), index=False)
    all_weekly_dep.to_csv(os.path.join(output_dir, 'Tables', 'district_weekly_departures.csv'), index=False)
    all_seasonal_dep.to_csv(os.path.join(output_dir, 'Tables', 'district_seasonal_departures.csv'), index=False)
    all_annual_dep.to_csv(os.path.join(output_dir, 'Tables', 'district_annual_departures.csv'), index=False)
    combined_enso_summary.to_csv(os.path.join(output_dir, 'Tables', 'combined_enso_summary.csv'), index=False)
    all_anova.to_csv(os.path.join(output_dir, 'Tables', 'anova_kruskal_results.csv'), index=False)
    all_corr_reg.to_csv(os.path.join(output_dir, 'Tables', 'correlation_regression_results.csv'), index=False)
    all_trends.to_csv(os.path.join(output_dir, 'Tables', 'mann_kendall_trends.csv'), index=False)
    
    # 6. Generate State-level Summaries & Visualizations
    logger.info("Generating state-level visualizations...")
    # Add state average plots
    state_monthly_dep = all_monthly_dep.groupby(['Year', 'Month']).agg({'Actual': 'mean', 'Normal': 'mean'}).reset_index()
    from departure import safe_pct_departure
    state_monthly_dep['Departure'] = safe_pct_departure(state_monthly_dep['Actual'].values, state_monthly_dep['Normal'].values)
    state_monthly_dep['District'] = 'State'
    state_monthly_dep['ENSO_Phase'] = state_monthly_dep['Year'].map(enso_map)
    
    state_weekly_dep = all_weekly_dep.groupby(['Year', 'SMW']).agg({'Actual': 'mean', 'Normal': 'mean'}).reset_index()
    state_weekly_dep['Departure'] = safe_pct_departure(state_weekly_dep['Actual'].values, state_weekly_dep['Normal'].values)
    state_weekly_dep['District'] = 'State'
    state_weekly_dep['ENSO_Phase'] = state_weekly_dep['Year'].map(enso_map)
    
    state_seasonal_dep = all_seasonal_dep.groupby(['Year', 'Season']).agg({'Actual': 'mean', 'Normal': 'mean'}).reset_index()
    state_seasonal_dep['Departure'] = safe_pct_departure(state_seasonal_dep['Actual'].values, state_seasonal_dep['Normal'].values)
    state_seasonal_dep['District'] = 'State'
    state_seasonal_dep['ENSO_Phase'] = state_seasonal_dep['Year'].map(enso_map)
    
    state_annual_dep = all_annual_dep.groupby('Year').agg({'Actual': 'mean', 'Normal': 'mean'}).reset_index()
    state_annual_dep['Departure'] = safe_pct_departure(state_annual_dep['Actual'].values, state_annual_dep['Normal'].values)
    state_annual_dep['District'] = 'State'
    state_annual_dep['ENSO_Phase'] = state_annual_dep['Year'].map(enso_map)
    
    # State-level summary objects
    from monthly import process_monthly_enso_summary
    from weekly import process_weekly_enso_summary
    from season import process_seasonal_enso_summary, process_annual_enso_summary
    
    state_m_summary = process_monthly_enso_summary(state_monthly_dep, enso_map)
    state_w_summary = process_weekly_enso_summary(state_weekly_dep, enso_map)
    state_s_summary = process_seasonal_enso_summary(state_seasonal_dep, enso_map)
    state_a_summary = process_annual_enso_summary(state_annual_dep, enso_map)

    # Plot State Average Graphics — run all 6 plots concurrently
    from graphs import (
        plot_monthly_bar_comparison, plot_monthly_anomaly_line,
        plot_monthly_boxplot_violin, plot_monthly_heatmap,
        plot_weekly_ribbon_anomaly, plot_seasonal_annual_boxplots
    )

    state_fig_dir = os.path.join(output_dir, 'Figures')

    state_plot_tasks = [
        (plot_monthly_bar_comparison,    (state_m_summary,   'State', state_fig_dir)),
        (plot_monthly_anomaly_line,       (state_monthly_dep, 'State', state_fig_dir)),
        (plot_monthly_boxplot_violin,     (state_monthly_dep, 'State', state_fig_dir)),
        (plot_monthly_heatmap,            (state_m_summary,   'State', state_fig_dir)),
        (plot_weekly_ribbon_anomaly,      (state_weekly_dep,  'State', state_fig_dir)),
        (plot_seasonal_annual_boxplots,   (state_seasonal_dep, state_annual_dep, 'State', state_fig_dir)),
    ]

    def _run_plot(fn_args):
        fn, a = fn_args
        fn(*a)

    logger.info("Generating state-level visualizations (parallel threads)...")
    with ThreadPoolExecutor(max_workers=6) as tex:
        plot_futs = [tex.submit(_run_plot, t) for t in state_plot_tasks]

        # While plots render, kick off spatial maps on the main thread
        # 7. Generate Spatial Maps
        from maps import generate_all_spatial_maps
        generate_all_spatial_maps(all_s_summary, all_m_summary, all_a_summary, output_dir)

        # Wait for all state plots to finish
        for fut in as_completed(plot_futs):
            exc = fut.exception()
            if exc:
                logger.warning(f"State plot task raised an exception: {exc}")

    # 8. Generate Reports — run all 3 report formats in parallel threads
    from report import generate_markdown_report, generate_docx_report, generate_pdf_report

    def _gen_markdown():
        try:
            generate_markdown_report(output_dir, state_s_summary, all_trends, all_anova)
        except Exception as e:
            logger.error(f"Failed to generate Markdown report: {e}")

    def _gen_docx():
        try:
            generate_docx_report(output_dir, state_s_summary, all_trends, all_anova)
        except Exception as e:
            logger.error(f"Failed to generate Word report: {e}")

    def _gen_pdf():
        try:
            generate_pdf_report(output_dir, state_s_summary, all_trends, all_anova)
        except Exception as e:
            logger.error(f"Failed to generate PDF report: {e}")

    logger.info("Generating reports in parallel (Markdown, Docx, PDF)...")
    with ThreadPoolExecutor(max_workers=3) as rex:
        r_futs = [rex.submit(f) for f in (_gen_markdown, _gen_docx, _gen_pdf)]
        for fut in as_completed(r_futs):
            exc = fut.exception()
            if exc:
                logger.warning(f"Report task raised: {exc}")
    
    # 9. Create Styled Excel File
    excel_path = os.path.join(output_dir, 'Excel', 'Rainfall_Analysis.xlsx')
    logger.info(f"Writing final styled Excel workbook to {excel_path}...")
    
    # Reorganize columns for standard sheets
    monthly_sheet_df = all_monthly_dep[['District', 'Year', 'Month', 'Actual', 'Normal', 'Departure']].copy()
    weekly_sheet_df = all_weekly_dep[['District', 'Year', 'SMW', 'Actual', 'Normal', 'Departure']].copy()
    sw_sheet_df = all_seasonal_dep[all_seasonal_dep['Season'] == 'SW Monsoon'][['District', 'Year', 'Actual', 'Normal', 'Departure']].copy()
    ne_sheet_df = all_seasonal_dep[all_seasonal_dep['Season'] == 'NE Monsoon'][['District', 'Year', 'Actual', 'Normal', 'Departure']].copy()
    annual_sheet_df = all_annual_dep[['District', 'Year', 'Actual', 'Normal', 'Departure']].copy()
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        write_styled_excel(writer, monthly_sheet_df, 'Monthly')
        write_styled_excel(writer, weekly_sheet_df, 'Weekly')
        write_styled_excel(writer, sw_sheet_df, 'SW Monsoon')
        write_styled_excel(writer, ne_sheet_df, 'NE Monsoon')
        write_styled_excel(writer, annual_sheet_df, 'Annual')
        write_styled_excel(writer, combined_enso_summary, 'ENSO Summary')
        write_styled_excel(writer, all_corr_reg, 'Statistics')
        write_styled_excel(writer, all_anova, 'ANOVA')
        write_styled_excel(writer, all_trends, 'Trend')
        
    logger.info("Excel workbook written successfully.")
    logger.info("ENSO rainfall analysis pipeline run completed successfully!")

if __name__ == '__main__':
    main()
